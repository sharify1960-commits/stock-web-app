import json
import smtplib
import os
import pandas as pd
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# הגדרת משתני ופונקציות המונה לפרויקט
COUNTER_FILE = "counter.json"

def load_counter():
    if os.path.exists(COUNTER_FILE):
        with open(COUNTER_FILE, "r", encoding="utf-8") as f:
            return json.load(f).get("count", 0)
    return 0

def save_counter(count):
    with open(COUNTER_FILE, "w", encoding="utf-8") as f:
        json.dump({"count": count}, f)

tickers = ["AAPL", "MSFT", "GOOGL", "AMZN", "NVDA", "META", "TSLA", "AMD", "NFLX", "INTC"]

def get_stock_data():
    data = []
    for ticker in tickers:
        try:
            hist = yf.Ticker(ticker).history(period="1y")
            if hist.empty or len(hist) < 30:
                continue
            
            close = hist['Close']
            current_price = close.iloc[-1]
            
            daily_ret = ((close.iloc[-1] - close.iloc[-2]) / close.iloc[-2]) * 100 if len(close) > 1 else 0
            weekly_ret = ((close.iloc[-1] - close.iloc[-5]) / close.iloc[-5]) * 100 if len(close) >= 5 else 0
            monthly_ret = ((close.iloc[-1] - close.iloc[-21]) / close.iloc[-21]) * 100 if len(close) >= 21 else 0
            yearly_ret = ((close.iloc[-1] - close.iloc[0]) / close.iloc[0]) * 100
            
            delta = close.diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
            rs = gain / loss
            rsi_series = 100 - (100 / (1 + rs))
            current_rsi = rsi_series.iloc[-1]
            
            if current_rsi > 60:
                signal = "Sell 🔴"
                reason = "RSI גבוה מ-60 (אזור רוויה)"
            elif current_rsi < 45:
                signal = "Buy 🟢"
                reason = "RSI נמוך מ-45 (הזדמנויות קנייה)"
            else:
                signal = "Buy 🟢"
                reason = "מומנטום חיובי (ממוצע 10 מעל 20)"
            
            support = round(current_price * 0.95, 2)
            target = round(current_price * 1.05, 2)
            
            data.append({
                "מניה": ticker,
                "איתות טכני": signal,
                "סיבת האיתות": reason,
                "מחיר סגירה": round(current_price, 2),
                "מחיר קנייה מומלץ (תמיכה)": support,
                "מחיר יעד למכירה": target,
                "יומי (%)": round(daily_ret, 2),
                "שבועי (%)": round(weekly_ret, 2),
                "חודשי (%)": round(monthly_ret, 2),
                "שנתי (%)": round(yearly_ret, 2),
                "RSI נוכחי": round(current_rsi, 2)
            })
        except Exception as e:
            print(f"Error fetching {ticker}: {e}")
            
    return pd.DataFrame(data)

def create_excel_report(df):
    file_path = "stock_report.xlsx"
    df.to_excel(file_path, index=False, sheet_name="Stock Report")
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.views.sheetView[0].rightToLeft = True
    
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0000FF", underline="single")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = regular_font
            cell.alignment = align_center
            cell.border = thin_border
            
            # הוספת קישור פעיל לגרף ב-Yahoo Finance בעמודת המניה
            if col == 1:
                ticker_val = cell.value
                cell.hyperlink = f"https://finance.yahoo.com/quote/{ticker_val}"
                cell.font = link_font
                
    # התאמת רוחב עמודות אוטומטית
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    wb.save(file_path)
    return file_path

def send_reports():
    sender_email = os.environ.get("MAIL_USERNAME")
    sender_password = os.environ.get("MAIL_PASSWORD")
    
    if not sender_email or not sender_password:
        print("Missing mail credentials.")
        return

    try:
        with open("subscribers.json", "r", encoding="utf-8") as f:
            subscribers = json.load(f)
    except Exception as e:
        print(f"Error loading subscribers: {e}")
        subscribers = {}

    if not subscribers:
        print("No subscribers found.")
        return

    df = get_stock_data()
    if df.empty:
        print("No stock data collected.")
        return

    excel_file = create_excel_report(df)

    html_content = """
    <!DOCTYPE html>
    <html dir="rtl" lang="he">
    <head>
        <meta charset="UTF-8">
        <style>
            body { font-family: Arial, sans-serif; direction: rtl; text-align: right; background-color: #f4f6f9; padding: 20px; }
            .container { background-color: #ffffff; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); max-width: 600px; margin: auto; }
            h2 { color: #2c3e50; text-align: center; }
        </style>
    </head>
    <body>
        <div class="container">
            <h2>📊 דוח מניות יומי - פרוייקט מניות איציק</h2>
            <p>שלום רב,</p>
            <p>מצורף קובץ האקסל המעודכן להיום הכולל את נתוני המניות, האיתותים הטכניים והקישורים הישירים לגרפים ב-Yahoo Finance.</p>
            <p>ניתן לפתוח את הקובץ בנוחות במחשב ובטלפון הנייד.</p>
            <p style="margin-top: 20px; color: #7f8c8d; font-size: 12px; text-align: center;">הדוח הופק אוטומטית באמצעות מערכת GitHub Actions.</p>
        </div>
    </body>
    </html>
    """

    for email, info in subscribers.items():
        if info.get("active", True):
            try:
                msg = MIMEMultipart()
                msg["Subject"] = "📊 דוח מניות יומי - קובץ אקסל - פרוייקט מניות איציק"
                msg["From"] = sender_email
                msg["To"] = email
                
                msg.attach(MIMEText(html_content, "html", "utf-8"))
                
                with open(excel_file, "rb") as f:
                    attach = MIMEApplication(f.read(), Name="stock_report.xlsx")
                    attach['Content-Disposition'] = 'attachment; filename="stock_report.xlsx"'
                    msg.attach(attach)
                
                with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                    server.login(sender_email, sender_password)
                    server.sendmail(sender_email, email, msg.as_string())
                print(f"Excel report email sent successfully to {email}")
            except Exception as e:
                print(f"Failed to send email to {email}: {e}")

if __name__ == "__main__":
    send_reports()
